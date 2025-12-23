#!/usr/bin/env python3
"""
Build 27x16 Policy Coverage Matrix Workbook

Creates a 4-tab Excel workbook showing which Henry Schein plans have
which policy coverage (FULL/LIMITED/NO) and BHG policy applicability.

Usage:
    python3 scripts/build-policy-matrix.py

Output:
    Henry_Schein_Policy_Coverage_Matrix.xlsx
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

# File paths
OUTPUT_DIR = Path(".")
OUTPUT_FILE = OUTPUT_DIR / "Henry_Schein_Policy_Coverage_Matrix.xlsx"
JSON_PLAN_FILE = Path("scripts/output/json-plan-analysis.json")
DRAFT_POLICIES_FILE = Path("scripts/output/draft-policies-summary.json")

# 16 Policy Areas (standardized)
POLICY_AREAS = [
    "Windfall/Large Deals",
    "Quota Management",
    "Territory Management",
    "Sales Crediting",
    "Clawback/Recovery",
    "SPIF Governance",
    "Termination/Final Pay",
    "New Hire/Onboarding",
    "Leave of Absence",
    "Payment Timing",
    "Compliance (409A, State Wage)",
    "Exceptions/Disputes",
    "Data/Systems/Controls",
    "Draws/Guarantees",
    "Mid-Period Changes",
    "International Requirements",
]

# Mapping of BHG DRAFT policies to policy areas they address
BHG_POLICY_MAPPING = {
    "Clawback And Recovery Policy": ["Clawback/Recovery", "Termination/Final Pay"],
    "Quota Management Policy": ["Quota Management", "Mid-Period Changes"],
    "Windfall Large Deal Policy": ["Windfall/Large Deals", "Exceptions/Disputes"],
    "Spif Governance Policy": ["SPIF Governance"],
    "Section 409A Compliance Policy": ["Compliance (409A, State Wage)", "Payment Timing", "Termination/Final Pay"],
    "State Wage Law Compliance Policy": ["Compliance (409A, State Wage)", "Payment Timing"],
}

# Colors
COLOR_FULL = "C6EFCE"  # Light green
COLOR_LIMITED = "FFEB9C"  # Light yellow
COLOR_NO = "FFC7CE"  # Light red
COLOR_HEADER = "4472C4"  # Blue
COLOR_ALT_ROW = "F2F2F2"  # Light gray


def create_tab1_coverage_summary(wb: Workbook, plans: List[Dict], policy_areas: List[str]):
    """Tab 1: Plan Policy Coverage Summary Matrix"""
    ws = wb.create_sheet("Plan Coverage Summary", 0)

    # Title
    ws['A1'] = "Henry Schein Compensation Plans - Policy Coverage Matrix"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:R1')

    # Metadata
    ws['A2'] = f"Total Plans: {len(plans)} | Policy Areas: {len(policy_areas)} | Generated: 2025-12"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:R2')

    # Header row
    row = 4
    ws[f'A{row}'] = "Plan Name"
    ws[f'A{row}'].font = Font(bold=True, color="FFFFFF")
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")

    for col_idx, policy in enumerate(policy_areas, start=2):
        cell = ws[f'{get_column_letter(col_idx)}{row}']
        cell.value = policy
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add Coverage % column
    ws[f'{get_column_letter(len(policy_areas) + 2)}{row}'] = "Coverage %"
    ws[f'{get_column_letter(len(policy_areas) + 2)}{row}'].font = Font(bold=True, color="FFFFFF")
    ws[f'{get_column_letter(len(policy_areas) + 2)}{row}'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws[f'{get_column_letter(len(policy_areas) + 2)}{row}'].alignment = Alignment(horizontal="center")

    # Data rows
    for plan_idx, plan in enumerate(plans, start=1):
        row = 4 + plan_idx
        plan_name = plan['planName']

        # Plan name
        ws[f'A{row}'] = plan_name
        ws[f'A{row}'].font = Font(size=10)
        if plan_idx % 2 == 0:
            ws[f'A{row}'].fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")

        # Policy coverage
        policy_coverage = plan.get('policyCoverage', {})
        for col_idx, policy_area in enumerate(policy_areas, start=2):
            cell = ws[f'{get_column_letter(col_idx)}{row}']

            # Get coverage level
            if policy_area in policy_coverage:
                coverage = policy_coverage[policy_area]['coverage']
            else:
                coverage = "NO"

            cell.value = coverage

            # Color coding
            if coverage == "FULL":
                cell.fill = PatternFill(start_color=COLOR_FULL, end_color=COLOR_FULL, fill_type="solid")
            elif coverage == "LIMITED":
                cell.fill = PatternFill(start_color=COLOR_LIMITED, end_color=COLOR_LIMITED, fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=COLOR_NO, end_color=COLOR_NO, fill_type="solid")

            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=9, bold=True)

        # Coverage percentage
        stats = plan.get('coverageStats', {})
        pct = stats.get('percentage', 0)
        pct_cell = ws[f'{get_column_letter(len(policy_areas) + 2)}{row}']
        pct_cell.value = f"{pct}%"
        pct_cell.alignment = Alignment(horizontal="center")
        pct_cell.font = Font(size=10, bold=True)
        if plan_idx % 2 == 0:
            pct_cell.fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 35
    for col_idx in range(2, len(policy_areas) + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Legend
    legend_row = row + 3
    ws[f'A{legend_row}'] = "Legend:"
    ws[f'A{legend_row}'].font = Font(bold=True)
    ws[f'B{legend_row}'] = "FULL"
    ws[f'B{legend_row}'].fill = PatternFill(start_color=COLOR_FULL, end_color=COLOR_FULL, fill_type="solid")
    ws[f'C{legend_row}'] = "Detailed enforceable policy with thresholds, workflows, SLAs"
    ws[f'D{legend_row}'] = "LIMITED"
    ws[f'D{legend_row}'].fill = PatternFill(start_color=COLOR_LIMITED, end_color=COLOR_LIMITED, fill_type="solid")
    ws[f'E{legend_row}'] = "Mentions policy area but lacks detail or clear process"
    ws[f'F{legend_row}'] = "NO"
    ws[f'F{legend_row}'].fill = PatternFill(start_color=COLOR_NO, end_color=COLOR_NO, fill_type="solid")
    ws[f'G{legend_row}'] = "Silent on policy area or only disclaimer language"

    print(f"✅ Tab 1 created: Plan Coverage Summary ({len(plans)} plans x {len(policy_areas)} policies)")


def create_tab2_gap_details(wb: Workbook, plans: List[Dict], policy_areas: List[str]):
    """Tab 2: Policy Gap Details"""
    ws = wb.create_sheet("Gap Details", 1)

    # Title
    ws['A1'] = "Policy Gap Analysis - NO and LIMITED Coverage Details"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    # Header
    headers = ["Plan Name", "Policy Area", "Current Coverage", "What's Missing", "BHG Policy That Addresses This", "Priority", "Risk Impact"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws[f'{get_column_letter(col_idx)}3']
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Extract gaps
    row = 4
    for plan in plans:
        plan_name = plan['planName']
        policy_coverage = plan.get('policyCoverage', {})

        for policy_area in policy_areas:
            if policy_area not in policy_coverage:
                coverage = "NO"
                details = "Policy area not addressed in plan documentation"
            else:
                coverage = policy_coverage[policy_area]['coverage']
                details = policy_coverage[policy_area].get('details', 'No details available')[:200]

            # Only include NO and LIMITED
            if coverage in ["NO", "LIMITED"]:
                ws[f'A{row}'] = plan_name
                ws[f'B{row}'] = policy_area
                ws[f'C{row}'] = coverage
                ws[f'D{row}'] = details

                # Find BHG policy
                bhg_policy = ""
                for policy, areas in BHG_POLICY_MAPPING.items():
                    if policy_area in areas:
                        bhg_policy = policy
                        break
                ws[f'E{row}'] = bhg_policy if bhg_policy else "Not addressed by BHG policies"

                # Priority (HIGH if NO, MEDIUM if LIMITED)
                priority = "HIGH" if coverage == "NO" else "MEDIUM"
                ws[f'F{row}'] = priority
                cell = ws[f'F{row}']
                if priority == "HIGH":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                # Risk impact
                if policy_area in ["Windfall/Large Deals", "Compliance (409A, State Wage)", "Clawback/Recovery"]:
                    risk = "CRITICAL"
                elif policy_area in ["Quota Management", "SPIF Governance", "Termination/Final Pay"]:
                    risk = "HIGH"
                else:
                    risk = "MEDIUM"
                ws[f'G{row}'] = risk

                # Styling
                for col in range(1, 8):
                    cell = ws[f'{get_column_letter(col)}{row}']
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.font = Font(size=9)

                row += 1

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12

    print(f"✅ Tab 2 created: Gap Details ({row - 4} gaps identified)")


def create_tab3_bhg_applicability(wb: Workbook, plans: List[Dict]):
    """Tab 3: BHG Policy Applicability"""
    ws = wb.create_sheet("BHG Policy Applicability", 2)

    # Title
    ws['A1'] = "BHG DRAFT Policy Applicability Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')

    # Header
    headers = ["BHG DRAFT Policy", "Policy Areas Addressed", "# Plans Needing This", "Plan Names", "Priority", "Implementation Complexity"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws[f'{get_column_letter(col_idx)}3']
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Analyze each BHG policy
    row = 4
    for bhg_policy, policy_areas_covered in BHG_POLICY_MAPPING.items():
        ws[f'A{row}'] = bhg_policy
        ws[f'B{row}'] = "\n".join(policy_areas_covered)

        # Count plans that need this policy (have NO or LIMITED in covered areas)
        plans_needing = []
        for plan in plans:
            policy_coverage = plan.get('policyCoverage', {})
            for area in policy_areas_covered:
                if area in policy_coverage:
                    coverage = policy_coverage[area]['coverage']
                    if coverage in ["NO", "LIMITED"]:
                        plans_needing.append(plan['planName'])
                        break

        ws[f'C{row}'] = len(plans_needing)
        ws[f'D{row}'] = "\n".join(plans_needing[:10]) + ("..." if len(plans_needing) > 10 else "")

        # Priority (MUST HAVE if >70% plans need, SHOULD HAVE if 40-70%, NICE TO HAVE if <40%)
        pct_needing = (len(plans_needing) / len(plans)) * 100
        if pct_needing > 70:
            priority = "MUST HAVE"
            color = "FFC7CE"
        elif pct_needing > 40:
            priority = "SHOULD HAVE"
            color = "FFEB9C"
        else:
            priority = "NICE TO HAVE"
            color = "C6EFCE"

        ws[f'E{row}'] = priority
        ws[f'E{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws[f'E{row}'].font = Font(bold=True)

        # Complexity (Windfall/Clawback = HIGH, others = MEDIUM)
        if any(keyword in bhg_policy.lower() for keyword in ["windfall", "clawback", "409a"]):
            complexity = "HIGH"
        else:
            complexity = "MEDIUM"
        ws[f'F{row}'] = complexity

        # Styling
        for col in range(1, 7):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=9)

        row += 1

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    print(f"✅ Tab 3 created: BHG Policy Applicability ({len(BHG_POLICY_MAPPING)} policies analyzed)")


def create_tab4_plan_details(wb: Workbook, plans: List[Dict]):
    """Tab 4: Plan Details Inventory"""
    ws = wb.create_sheet("Plan Details", 3)

    # Title
    ws['A1'] = "Henry Schein Compensation Plan Inventory"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')

    # Header
    headers = ["Plan Name", "Business Unit", "Plan Type", "Coverage %", "Full Policies", "Limited Policies", "No Policies", "Source File"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws[f'{get_column_letter(col_idx)}3']
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for plan_idx, plan in enumerate(plans, start=1):
        row = 3 + plan_idx

        ws[f'A{row}'] = plan['planName']
        ws[f'B{row}'] = plan.get('businessUnit', 'Unknown')
        ws[f'C{row}'] = plan.get('planType', 'Unknown')

        stats = plan.get('coverageStats', {})
        ws[f'D{row}'] = f"{stats.get('percentage', 0)}%"
        ws[f'E{row}'] = stats.get('full', 0)
        ws[f'F{row}'] = stats.get('limited', 0)
        ws[f'G{row}'] = stats.get('no', 0)
        ws[f'H{row}'] = plan.get('sourceFile', '')

        # Styling
        for col in range(1, 9):
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=9)
            if plan_idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 40

    print(f"✅ Tab 4 created: Plan Details ({len(plans)} plans)")


def main():
    print("🚀 Building 27x16 Policy Coverage Matrix workbook...")
    print(f"📂 Output: {OUTPUT_FILE}\n")

    # Load JSON data
    if not JSON_PLAN_FILE.exists():
        print(f"❌ Error: JSON plan file not found: {JSON_PLAN_FILE}")
        return 1

    with open(JSON_PLAN_FILE, 'r') as f:
        plan_data = json.load(f)

    plans = plan_data['plans']
    policy_areas = plan_data['metadata']['standardPolicyAreas']

    print(f"📊 Loaded {len(plans)} plans and {len(policy_areas)} policy areas\n")

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create 4 tabs
    create_tab1_coverage_summary(wb, plans, policy_areas)
    create_tab2_gap_details(wb, plans, policy_areas)
    create_tab3_bhg_applicability(wb, plans)
    create_tab4_plan_details(wb, plans)

    # Save workbook
    wb.save(OUTPUT_FILE)
    print(f"\n✅ Workbook saved: {OUTPUT_FILE}")
    print(f"📊 4 tabs created:")
    print(f"   1. Plan Coverage Summary (27x16 matrix)")
    print(f"   2. Gap Details (NO and LIMITED policies)")
    print(f"   3. BHG Policy Applicability (6 DRAFT policies)")
    print(f"   4. Plan Details (plan inventory)")
    print()
    print("🎉 Policy matrix complete!")
    return 0


if __name__ == '__main__':
    exit(main())
