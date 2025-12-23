#!/usr/bin/env python3
"""
Enhance Henry Schein Deliverables Mapping

Reads existing CSV and enhances it with:
- File existence validation
- File sizes and last modified dates
- Plan applicability mapping
- Policy coverage levels
- Risk mitigation values

Usage:
    python3 scripts/enhance-mapping.py

Output:
    Henry_Schein_Deliverables_Mapping_CORRECTED.xlsx
"""

import csv
import json
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any

# File paths
ARCHIVE_ROOT = Path("/Users/toddlebaron/dev__archive_20251219_1518/clients/HenrySchien")
DELIVERY_PKG = ARCHIVE_ROOT / "CLIENT_DELIVERY_PACKAGE"
INPUT_CSV = Path("Henry_Schein_Readout_Deliverables_Mapping.csv")
OUTPUT_FILE = Path("Henry_Schein_Deliverables_Mapping_CORRECTED.xlsx")
JSON_PLAN_FILE = Path("scripts/output/json-plan-analysis.json")

# Risk exposure by deliverable type/priority
RISK_VALUES = {
    "CRITICAL": {
        "Policy": 500000,
        "Framework": 300000,
        "Assessment": 200000,
        "Procedure": 100000,
    },
    "HIGH": {
        "Policy": 250000,
        "Framework": 150000,
        "Assessment": 100000,
        "Procedure": 50000,
    },
    "MEDIUM": {
        "Policy": 100000,
        "Framework": 50000,
        "Assessment": 25000,
        "Procedure": 10000,
    },
    "IMMEDIATE": {  # Same as CRITICAL
        "Policy": 500000,
        "Framework": 300000,
        "Assessment": 200000,
        "Procedure": 100000,
    },
}

# Policy deliverable to policy area mapping
DELIVERABLE_TO_POLICY_AREA = {
    "CLAWBACK_AND_RECOVERY_POLICY": "Clawback/Recovery",
    "QUOTA_MANAGEMENT_POLICY": "Quota Management",
    "WINDFALL_LARGE_DEAL_POLICY": "Windfall/Large Deals",
    "SPIF_GOVERNANCE_POLICY": "SPIF Governance",
    "SECTION_409A_COMPLIANCE_POLICY": "Compliance (409A, State Wage)",
    "STATE_WAGE_LAW_COMPLIANCE_POLICY": "Compliance (409A, State Wage)",
    "SALES_CREDITING_POLICY": "Sales Crediting",
    "TERMINATION_POLICY": "Termination/Final Pay",
    "PAYMENT_TIMING_POLICY": "Payment Timing",
    "MID_PERIOD_CHANGE_POLICY": "Mid-Period Changes",
    "LEAVE_OF_ABSENCE_POLICY": "Leave of Absence",
    "DRAWS_AND_GUARANTEES_POLICY": "Draws/Guarantees",
    "DISPUTE_RESOLUTION": "Exceptions/Disputes",
    "EXCEPTION_REQUEST": "Exceptions/Disputes",
}


def check_file_exists(file_path: str) -> tuple[bool, str, str, int]:
    """
    Check if deliverable file exists.

    Returns:
        (exists, status, actual_path, file_size)
    """
    if not file_path:
        return False, "NO PATH", "", 0

    # Try full path
    full_path = DELIVERY_PKG / file_path

    if full_path.exists():
        stat = full_path.stat()
        return True, "YES", str(full_path), stat.st_size
    else:
        # Check if it's a DRAFT
        if "DRAFT" in file_path:
            return False, "DRAFT", str(full_path), 0
        else:
            return False, "MISSING", str(full_path), 0


def format_file_size(size_bytes: int) -> str:
    """Format file size in human-readable format."""
    if size_bytes == 0:
        return "0 KB"
    elif size_bytes < 1024:
        return f"{size_bytes} B"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.1f} KB"
    else:
        return f"{size_bytes / (1024 * 1024):.1f} MB"


def get_policy_area_from_deliverable(file_path: str) -> str:
    """Extract policy area from deliverable file path."""
    for key, policy_area in DELIVERABLE_TO_POLICY_AREA.items():
        if key in file_path.upper():
            return policy_area
    return "Unknown"


def find_applicable_plans(policy_area: str, plan_data: List[Dict]) -> tuple[List[str], str]:
    """
    Find plans that need this policy (have NO or LIMITED coverage).

    Returns:
        (list of plan names, coverage summary)
    """
    if policy_area == "Unknown":
        return [], "N/A"

    plans_needing = []
    for plan in plan_data:
        policy_coverage = plan.get('policyCoverage', {})
        if policy_area in policy_coverage:
            coverage = policy_coverage[policy_area]['coverage']
            if coverage in ["NO", "LIMITED"]:
                plans_needing.append(plan['planName'])

    if not plans_needing:
        summary = "All plans have full coverage"
    elif len(plans_needing) < 5:
        summary = f"{len(plans_needing)} plans need this"
    else:
        summary = f"{len(plans_needing)} plans need this"

    return plans_needing, summary


def calculate_risk_mitigated(priority: str, deliverable_type: str, plans_count: int) -> int:
    """Calculate risk mitigated based on priority, type, and # plans affected."""
    base_risk = RISK_VALUES.get(priority, RISK_VALUES["MEDIUM"]).get(deliverable_type, 10000)

    # Scale by number of plans affected (if applicable)
    if plans_count > 0:
        # Each plan adds 10% to base risk
        multiplier = 1 + (plans_count * 0.1)
        return int(base_risk * multiplier)
    else:
        return base_risk


def main():
    print("🚀 Enhancing Henry Schein deliverables mapping...")
    print(f"📂 Input: {INPUT_CSV}")
    print(f"📂 Output: {OUTPUT_FILE}\n")

    # Load plan data
    if JSON_PLAN_FILE.exists():
        with open(JSON_PLAN_FILE, 'r') as f:
            plan_data_json = json.load(f)
            plan_data = plan_data_json['plans']
    else:
        print("⚠️  Plan data not found, continuing without plan mapping")
        plan_data = []

    # Read existing CSV
    if not INPUT_CSV.exists():
        print(f"❌ Input CSV not found: {INPUT_CSV}")
        return 1

    print(f"📖 Reading existing mapping CSV...")
    with open(INPUT_CSV, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    print(f"✅ Loaded {len(rows)} rows\n")

    # Enhance each row
    print("🔍 Validating and enhancing deliverables...")
    enhanced_rows = []

    stats = {
        'exists': 0,
        'missing': 0,
        'draft': 0,
    }

    for i, row in enumerate(rows, 1):
        if i % 10 == 0:
            print(f"   Processing row {i}/{len(rows)}...")

        file_path = row.get('Deliverable File Path', '')
        priority = row.get('Priority', 'MEDIUM')
        deliverable_type = row.get('Deliverable Type', 'Unknown')

        # Check if file exists
        exists, status, actual_path, file_size = check_file_exists(file_path)

        # Update stats
        if status == "YES":
            stats['exists'] += 1
        elif status == "DRAFT":
            stats['draft'] += 1
        else:
            stats['missing'] += 1

        # Get policy area
        policy_area = get_policy_area_from_deliverable(file_path)

        # Find applicable plans
        applicable_plans, plans_summary = find_applicable_plans(policy_area, plan_data)

        # Calculate risk mitigated
        risk_mitigated = calculate_risk_mitigated(priority, deliverable_type, len(applicable_plans))

        # Enhanced row
        enhanced_row = {
            **row,  # Keep all original columns
            'Deliverable Exists?': status,
            'File Size': format_file_size(file_size),
            'Actual File Path': actual_path if exists else 'NOT FOUND',
            'Policy Area': policy_area,
            'Applies to Plans': plans_summary,
            'Plan Names': ', '.join(applicable_plans[:5]) + ('...' if len(applicable_plans) > 5 else ''),
            'Plans Count': len(applicable_plans),
            'Risk Mitigated ($)': f"${risk_mitigated:,}",
            'Validation Notes': f"Verified {datetime.now().strftime('%Y-%m-%d')}" if exists else "File not found in delivery package",
        }

        enhanced_rows.append(enhanced_row)

    print(f"✅ Enhanced {len(enhanced_rows)} rows\n")

    # Create Excel workbook
    print("📊 Creating Excel workbook...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliverables Mapping"

    # Title
    ws['A1'] = "Henry Schein Deliverables Mapping - CORRECTED & ENHANCED"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:Q1')

    # Stats
    ws['A2'] = f"Total: {len(enhanced_rows)} | Exists: {stats['exists']} | Missing: {stats['missing']} | Draft: {stats['draft']} | Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:Q2')

    # Headers
    if enhanced_rows:
        headers = list(enhanced_rows[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws[f'{get_column_letter(col_idx)}4']
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for row_idx, row_data in enumerate(enhanced_rows, start=5):
            for col_idx, (key, value) in enumerate(row_data.items(), start=1):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Color code by existence status
                if key == 'Deliverable Exists?':
                    if value == "YES":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "DRAFT":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Column widths
    col_widths = {
        'A': 20,  # Category
        'B': 35,  # Readout Item
        'C': 25,  # PPT Reference
        'D': 40,  # Finding/Gap
        'E': 10,  # Priority
        'F': 15,  # Deliverable Type
        'G': 50,  # Deliverable File Path
        'H': 20,  # Implementation Phase
        'I': 15,  # Status
        'J': 15,  # Deliverable Exists?
        'K': 10,  # File Size
        'L': 50,  # Actual File Path
        'M': 20,  # Policy Area
        'N': 20,  # Applies to Plans
        'O': 40,  # Plan Names
        'P': 10,  # Plans Count
        'Q': 15,  # Risk Mitigated
        'R': 30,  # Validation Notes
    }

    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook
    wb.save(OUTPUT_FILE)

    print(f"✅ Workbook saved: {OUTPUT_FILE}\n")

    # Summary
    print("📊 Summary:")
    print("=" * 90)
    print(f"   Total Deliverables:    {len(enhanced_rows)}")
    print(f"   ✅ Exists:             {stats['exists']}")
    print(f"   📝 Draft:              {stats['draft']}")
    print(f"   ❌ Missing:            {stats['missing']}")
    print("=" * 90)
    print()

    # Top risk items
    print("🔴 Top Risk Mitigation Items:")
    print("=" * 90)
    sorted_rows = sorted(enhanced_rows, key=lambda x: int(x['Risk Mitigated ($)'].replace('$', '').replace(',', '')), reverse=True)
    for row in sorted_rows[:10]:
        name = row['Readout Item'][:40]
        risk = row['Risk Mitigated ($)']
        status = row['Deliverable Exists?']
        print(f"   {name:42} | {risk:15} | {status}")
    print("=" * 90)
    print()

    print("🎉 Mapping enhancement complete!")
    return 0


if __name__ == '__main__':
    exit(main())
