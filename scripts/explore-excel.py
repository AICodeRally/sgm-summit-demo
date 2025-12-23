#!/usr/bin/env python3
"""Quick script to explore Excel workbook structure"""

import openpyxl
from pathlib import Path

EXCEL_FILE = Path("/Users/toddlebaron/dev__archive_20251219_1518/clients/HenrySchien/Analysis/Comp Analysis/workbooks/master/BHG_01_HS_Comp_Plan_Analysis_FINAL.xlsx")

wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

# Explore key sheets
sheets_to_examine = [
    "12) Plan Coverage Matrix",
    "1) Executive Summary",
    "00_Overview"
]

for sheet_name in sheets_to_examine:
    if sheet_name in wb.sheetnames:
        print(f"\n{'='*80}")
        print(f"SHEET: {sheet_name}")
        print(f"{'='*80}")
        sheet = wb[sheet_name]

        # Print first 20 rows, first 10 columns
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True), 1):
            if any(row):  # Skip completely empty rows
                row_str = " | ".join([str(cell)[:30] if cell else "" for cell in row])
                print(f"Row {row_idx:2}: {row_str}")
