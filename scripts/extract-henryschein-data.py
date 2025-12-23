#!/usr/bin/env python3
"""
Henry Schein Data Extraction Script

Extracts plan inventory, risk scores, participant counts, and spend data
from the BHG master Excel workbook for gap analysis.

Usage:
    python3 scripts/extract-henryschein-data.py

Output:
    scripts/output/henryschein-plan-data.json

Requirements:
    pip install openpyxl
"""

import json
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# File paths
ARCHIVE_ROOT = Path("/Users/toddlebaron/dev__archive_20251219_1518/clients/HenrySchien")
EXCEL_FILE = ARCHIVE_ROOT / "Analysis/Comp Analysis/workbooks/master/BHG_01_HS_Comp_Plan_Analysis_FINAL.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_FILE = OUTPUT_DIR / "henryschein-plan-data.json"

# Division mapping
DIVISION_MAPPING = {
    'Medical': 'Medical',
    'Dental': 'Dental',
    'HSPG': 'HSPG',
    'Specialty': 'Specialty',
    'Corporate': 'Corporate',
}


def extract_plan_coverage_matrix(wb) -> tuple[List[Dict[str, Any]], List[str]]:
    """
    Extract plan coverage matrix from the "12) Plan Coverage Matrix" sheet.

    Returns:
        - plans: List of plan dictionaries with policy coverage
        - policy_areas: List of policy area names
    """
    sheet_name = "12) Plan Coverage Matrix"
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet '{sheet_name}' not found!")
        return [], []

    sheet = wb[sheet_name]
    print(f"✅ Found sheet: {sheet_name}")

    # Row 1 has plan names starting from column 3 (C)
    # Columns: Policy Area | Item | Plan1 | Plan2 | ...
    header_row = list(sheet[1])
    plan_columns = []
    plan_names = []

    for idx, cell in enumerate(header_row[2:], start=2):  # Skip first 2 columns
        if cell.value and str(cell.value).strip():
            plan_name = str(cell.value).strip()
            # Skip header columns like "Policy Area" or empty
            if plan_name and plan_name not in ['Policy Area', 'Item']:
                plan_columns.append(idx)
                plan_names.append(plan_name)

    print(f"📋 Found {len(plan_names)} plans: {plan_names[:5]}...")

    # Extract policy areas and their coverage
    # Rows 2-16 contain policy areas
    policy_areas = []
    plans = {name: {'planName': name, 'policyCoverage': {}} for name in plan_names}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=16, values_only=True), start=2):
        if not row or len(row) < 2:
            continue

        policy_area = str(row[0]).strip() if row[0] else None
        policy_item = str(row[1]).strip() if row[1] else None

        if not policy_area or policy_area == '% Coverage by Plan':
            continue

        # Use policy item as the detailed policy name, fallback to area
        policy_name = policy_item if policy_item else policy_area
        policy_areas.append(policy_name)

        # Extract coverage for each plan
        for plan_idx, col_idx in enumerate(plan_columns):
            plan_name = plan_names[plan_idx]
            coverage_value = str(row[col_idx]).strip().upper() if col_idx < len(row) and row[col_idx] else 'NO'

            # Map to our standard: FULL, LIMITED, NO
            if coverage_value == 'YES':
                coverage = 'FULL'
            elif coverage_value == 'PARTIAL':
                coverage = 'LIMITED'
            else:
                coverage = 'NO'

            plans[plan_name]['policyCoverage'][policy_name] = coverage

    # Convert dict to list
    plans_list = list(plans.values())

    print(f"✅ Extracted {len(policy_areas)} policy areas and coverage for {len(plans_list)} plans")

    return plans_list, policy_areas


def extract_executive_summary(wb, plan_names: List[str]) -> Dict[str, Dict[str, Any]]:
    """
    Extract plan details from Executive Summary sheet.

    Returns:
        Dict mapping plan name to plan details (division, risk, etc.)
    """
    sheet_name = "1) Executive Summary"
    if sheet_name not in wb.sheetnames:
        print(f"⚠️  Sheet '{sheet_name}' not found!")
        return {}

    sheet = wb[sheet_name]
    print(f"✅ Found sheet: {sheet_name}")

    # Row 5 has plan names
    # Look for specific attributes in rows below
    plan_details = {}

    # Read all data
    all_rows = list(sheet.iter_rows(min_row=1, max_row=30, values_only=True))

    # Find plan name row (should be row 5, index 4)
    plan_row_idx = None
    for idx, row in enumerate(all_rows):
        if row and len(row) > 1 and str(row[0]).strip() == 'Attribute':
            plan_row_idx = idx
            break

    if not plan_row_idx:
        print("⚠️  Could not find plan names in Executive Summary")
        return {}

    # Get plan names from that row (columns 1+)
    plan_row = all_rows[plan_row_idx]
    exec_plan_names = []
    for cell in plan_row[1:]:
        if cell and str(cell).strip():
            exec_plan_names.append(str(cell).strip())

    # Extract attributes for each plan
    attribute_map = {
        'Plan Type': 'planType',
        'Geography': 'geography',
        'Effective Date': 'effectiveDate',
        'Business Unit': 'businessUnit',
        'Overall Legal Risk': 'legalRisk',
        'Financial Liability Risk': 'financialRisk',
    }

    for col_idx, plan_name in enumerate(exec_plan_names, start=1):
        plan_details[plan_name] = {'planName': plan_name}

        # Scan rows for attributes
        for row_idx in range(plan_row_idx + 1, len(all_rows)):
            row = all_rows[row_idx]
            if not row or len(row) <= col_idx:
                continue

            attr_name = str(row[0]).strip() if row[0] else None
            attr_value = str(row[col_idx]).strip() if row[col_idx] else None

            if attr_name in attribute_map and attr_value:
                plan_details[plan_name][attribute_map[attr_name]] = attr_value

    print(f"✅ Extracted details for {len(plan_details)} plans from Executive Summary")

    return plan_details


def main():
    print("🚀 Starting Henry Schein data extraction...")
    print(f"📂 Source: {EXCEL_FILE}")
    print(f"📂 Output: {OUTPUT_FILE}\n")

    # Check if file exists
    if not EXCEL_FILE.exists():
        print(f"❌ Error: Excel file not found at {EXCEL_FILE}")
        print(f"   Please ensure the archive directory is accessible.")
        return 1

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load workbook
    print(f"📖 Loading workbook...")
    wb = load_workbook(EXCEL_FILE, data_only=True)
    print(f"✅ Workbook loaded. Found {len(wb.sheetnames)} sheets:")
    for idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"   {idx}. {sheet_name}")
    print()

    # Extract plan coverage matrix
    print("📊 Extracting plan coverage matrix...")
    plans, policy_areas = extract_plan_coverage_matrix(wb)
    print()

    if not plans:
        print("❌ No plans extracted from coverage matrix!")
        return 1

    # Extract executive summary details
    print("📊 Extracting executive summary...")
    plan_details = extract_executive_summary(wb, [p['planName'] for p in plans])
    print()

    # Merge plan details with coverage data
    print("🔗 Merging plan data...")
    for plan in plans:
        plan_name = plan['planName']
        if plan_name in plan_details:
            plan.update(plan_details[plan_name])
    print()

    # Calculate coverage statistics
    print("📊 Calculating coverage statistics...")
    coverage_stats = {
        'totalPolicyCoverage': {},
        'planCoveragePercentages': {},
    }

    # Calculate per-policy coverage across all plans
    for policy in policy_areas:
        coverage_stats['totalPolicyCoverage'][policy] = {
            'FULL': 0,
            'LIMITED': 0,
            'NO': 0,
        }
        for plan in plans:
            coverage = plan['policyCoverage'].get(policy, 'NO')
            coverage_stats['totalPolicyCoverage'][policy][coverage] += 1

    # Calculate per-plan coverage percentage
    for plan in plans:
        coverage = plan['policyCoverage']
        full = sum(1 for v in coverage.values() if v == 'FULL')
        limited = sum(1 for v in coverage.values() if v == 'LIMITED')
        total = len(policy_areas)
        percentage = round((full + 0.5 * limited) / total * 100, 1) if total > 0 else 0
        coverage_stats['planCoveragePercentages'][plan['planName']] = {
            'full': full,
            'limited': limited,
            'no': total - full - limited,
            'percentage': percentage,
        }

    # Compile output
    output_data = {
        'metadata': {
            'source': str(EXCEL_FILE),
            'extractionDate': str(Path(__file__).stat().st_mtime),
            'totalPlans': len(plans),
            'totalPolicyAreas': len(policy_areas),
        },
        'policyAreas': policy_areas,
        'plans': plans,
        'coverageStats': coverage_stats,
    }

    # Write to JSON
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)

    print(f"✅ Data extracted successfully!")
    print(f"📊 Summary:")
    print(f"   Total Plans: {len(plans)}")
    print(f"   Policy Areas: {len(policy_areas)}")
    print(f"   Output File: {OUTPUT_FILE}")
    print()

    # Display top/bottom plans by coverage
    print("📊 Plan Coverage Summary:")
    print("═" * 90)
    sorted_plans = sorted(
        coverage_stats['planCoveragePercentages'].items(),
        key=lambda x: x[1]['percentage'],
        reverse=True
    )
    for plan_name, stats in sorted_plans[:10]:  # Top 10
        full = stats['full']
        limited = stats['limited']
        no = stats['no']
        pct = stats['percentage']
        print(f"   {plan_name:30} | Full: {full:2} | Limited: {limited:2} | No: {no:2} | Coverage: {pct:5.1f}%")
    print("═" * 90)
    print()

    print("🎉 Extraction complete!")
    return 0


if __name__ == '__main__':
    exit(main())
